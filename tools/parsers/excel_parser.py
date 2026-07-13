from __future__ import annotations

import os
from collections.abc import Callable
from typing import Any

from tools.parsers.csv_parser import _apply_hooks


def parse_excel(
    filepath: str,
    sheet_name: str | int = 0,
    header_row: int = 1,
    row_filter: Callable[[dict[str, Any]], bool] | None = None,
    row_transform: Callable[[dict[str, Any]], dict[str, Any]] | None = None,
) -> tuple[list[str], list[list[Any]], str | None]:
    if not os.path.isfile(filepath):
        return [], [], f"File not found: {filepath}"

    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], [], "openpyxl is not installed. Run: uv sync"

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if isinstance(sheet_name, int):
            if sheet_name < 0 or sheet_name >= len(wb.worksheets):
                wb.close()
                return [], [], f"Sheet index {sheet_name} out of range"
            ws = wb.worksheets[sheet_name]
        else:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return [], [], f"Sheet '{sheet_name}' not found"
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(min_row=header_row, values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            headers = []
        raw_rows = [list(row) for row in rows_iter]
        wb.close()

        rows = _apply_hooks(raw_rows, headers, row_filter, row_transform)
        return headers, rows, None
    except Exception as e:
        return [], [], str(e)


def count_excel_rows(filepath: str, sheet_name: str | int = 0) -> int | None:
    if not filepath or not os.path.isfile(filepath):
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if isinstance(sheet_name, int):
            if sheet_name < 0 or sheet_name >= len(wb.worksheets):
                wb.close()
                return None
            ws = wb.worksheets[sheet_name]
        else:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return None
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            next(rows_iter)
        except StopIteration:
            wb.close()
            return 0
        count = sum(1 for _ in rows_iter)
        wb.close()
        return count
    except Exception:
        return None
