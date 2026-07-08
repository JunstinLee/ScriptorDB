from __future__ import annotations

import os
from typing import Any

from pydantic_ai import RunContext

from config.settings import Settings
from tools.errors import _to_tool_error
from tools.tool_result import ToolErrorInfo, ToolResult


def export_excel(
    ctx: RunContext[Settings],
    filepath: str,
    sheets: dict[str, dict],
) -> ToolResult:
    try:
        from openpyxl import Workbook
    except ImportError:
        return ToolResult(
            success=False,
            error=ToolErrorInfo(
                category="parameter_error",
                message="openpyxl is not installed. Run: uv sync",
            ),
        )

    try:
        wb = Workbook()
        ws = wb.active
        for sheet_name, sheet_data in sheets.items():
            if ws is not None and ws.title == "Sheet":
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            columns = sheet_data.get("columns", [])
            data = sheet_data.get("data", [])
            if columns:
                ws.append(columns)
            for row in data:
                ws.append(row)

        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        wb.save(filepath)

        total_rows = sum(len(s.get("data", [])) for s in sheets.values())
        abs_path = os.path.abspath(filepath)
        return ToolResult(
            success=True,
            output=f"Exported {len(sheets)} sheet{'s' if len(sheets) != 1 else ''}, {total_rows} row{'s' if total_rows != 1 else ''} to {os.path.basename(filepath)}",
            data={"file": abs_path, "sheets": len(sheets), "total_rows": total_rows},
        )
    except Exception as e:
        return _to_tool_error(e)


def read_excel(
    ctx: RunContext[Settings],
    filepath: str,
    sheet_name: str | int = 0,
    header_row: int = 1,
    preview_rows: int = 10,
    return_full: bool = False,
    max_rows: int | None = None,
) -> ToolResult:
    if not os.path.isfile(filepath):
        return ToolResult(
            success=False,
            error=ToolErrorInfo(
                category="resource_not_found",
                message=f"File not found: {filepath}",
            ),
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        return ToolResult(
            success=False,
            error=ToolErrorInfo(
                category="parameter_error",
                message="openpyxl is not installed. Run: uv sync",
            ),
        )

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if isinstance(sheet_name, int):
            if sheet_name < 0 or sheet_name >= len(wb.worksheets):
                return ToolResult(
                    success=False,
                    error=ToolErrorInfo(
                        category="parameter_error",
                        message=f"Sheet index {sheet_name} out of range",
                    ),
                )
            ws = wb.worksheets[sheet_name]
        else:
            if sheet_name not in wb.sheetnames:
                return ToolResult(
                    success=False,
                    error=ToolErrorInfo(
                        category="parameter_error",
                        message=f"Sheet '{sheet_name}' not found",
                    ),
                )
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(min_row=header_row, values_only=True)
        try:
            headers = list(next(rows_iter))
        except StopIteration:
            headers = []

        preview: list[list[Any | None]] = []
        rows_data: list[list[Any | None]] = []
        row_count = 0
        truncated = False
        for row in rows_iter:
            row_count += 1
            row_values = [str(v) if v is not None else None for v in row]
            if len(preview) < preview_rows:
                preview.append(row_values)
            if return_full:
                if max_rows is None or len(rows_data) < max_rows:
                    rows_data.append(row_values)
                else:
                    truncated = True

        wb.close()

        data: dict[str, Any] = {
            "file": filepath,
            "sheet": ws.title,
            "columns": headers,
            "rows": row_count,
            "preview": preview,
        }
        output = (
            f"Read {os.path.basename(filepath)}: sheet {ws.title}, {row_count} row"
            f"{'s' if row_count != 1 else ''}, {len(headers)} column"
            f"{'s' if len(headers) != 1 else ''}"
        )
        if return_full:
            data["data"] = rows_data
            data["truncated"] = truncated
            if truncated:
                output += f" (returned {len(rows_data)} of {row_count} rows)"

        return ToolResult(success=True, output=output, data=data)
    except Exception as e:
        return _to_tool_error(e)
