from __future__ import annotations

import csv
import os
from collections.abc import Callable
from typing import Any

from pydantic_ai import RunContext
from sqlalchemy import text

from config.settings import Settings
from tools.db_repository import DatabaseRepository
from tools.errors import _to_tool_error
from tools.parsers.csv_parser import _apply_hooks, parse_csv
from tools.parsers.excel_parser import parse_excel
from tools.schema_helpers import (
    create_table_from_headers,
    get_pk_columns,
    quote_identifier,
    table_exists,
    unique_table_name,
)
from tools.tool_decorators import db_tool
from tools.tool_result import ToolErrorInfo, ToolResult
from tools.validators import validate_import_args


def _build_insert_sql(table_name: str, headers: list[str]) -> str:
    cols = [quote_identifier(header) for header in headers]
    placeholders = [f":p{i}" for i in range(len(headers))]
    return (
        f"INSERT INTO {quote_identifier(table_name)} "
        f"({', '.join(cols)}) VALUES ({', '.join(placeholders)})"
    )


def _insert_batches(
    conn,
    table_name: str,
    headers: list[str],
    rows: list[list[Any]],
    batch_size: int,
    capture_rows: bool = False,
) -> tuple[int, list[dict] | None]:
    sql = _build_insert_sql(table_name, headers)
    total = 0
    inserted_rows: list[dict] = []
    pk_cols = get_pk_columns(conn, table_name) if capture_rows else []
    dialect = conn.dialect.name

    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        params = [
            {f"p{j}": row[j] for j in range(len(headers))}
            for row in batch
        ]

        if capture_rows and pk_cols:
            if dialect == "sqlite":
                returning_sql = sql.rstrip(";") + " RETURNING *"
                result = conn.execute(text(returning_sql), params)
                batch_rows = [dict(row._mapping) for row in result.fetchall()]
                inserted_rows.extend(batch_rows)
                total += len(batch_rows)
            elif dialect == "mysql":
                conn.execute(text(sql), params)
                total += len(batch)
                count = len(batch)
                last_id_result = conn.execute(text("SELECT LAST_INSERT_ID()"))
                last_id = last_id_result.scalar()
                if last_id is not None and pk_cols:
                    first_id = last_id - count + 1
                    pk_col = pk_cols[0]
                    select_sql = (
                        f"SELECT * FROM {quote_identifier(table_name)} "
                        f"WHERE {quote_identifier(pk_col)} BETWEEN :first AND :last"
                    )
                    select_result = conn.execute(
                        text(select_sql), {"first": first_id, "last": last_id}
                    )
                    batch_rows = [dict(row._mapping) for row in select_result.fetchall()]
                    inserted_rows.extend(batch_rows)
            else:
                conn.execute(text(sql), params)
                total += len(batch)
        else:
            conn.execute(text(sql), params)
            total += len(batch)

    return total, inserted_rows if capture_rows else None


def _build_undo_entries_for_inserted_rows(
    conn,
    table_name: str,
    inserted_rows: list[dict],
) -> list[tuple[str, dict]]:
    pk_cols = get_pk_columns(conn, table_name)
    if not pk_cols or not inserted_rows:
        return []

    undo_entries: list[tuple[str, dict]] = []
    for row in inserted_rows:
        pk_conditions = " AND ".join(
            f"{quote_identifier(col)} = :undo_{col}" for col in pk_cols
        )
        undo_sql = f"DELETE FROM {quote_identifier(table_name)} WHERE {pk_conditions}"
        undo_params = {f"undo_{col}": row[col] for col in pk_cols}
        undo_entries.append((undo_sql, undo_params))
    return undo_entries


def _import_rows_to_db(
    ctx: RunContext[Settings],
    table_name: str,
    headers: list[str],
    rows: list[list[Any]],
    if_exists: str,
    batch_size: int,
) -> ToolResult:
    repo = DatabaseRepository(ctx.deps.db_url, ctx.deps.workspace_id or "")
    try:
        with repo.session() as conn:
            exists = table_exists(conn, table_name)
            if exists:
                if if_exists == "fail":
                    return ToolResult(
                        success=False,
                        error=ToolErrorInfo(
                            category="parameter_error",
                            message=f"Table '{table_name}' already exists",
                        ),
                    )
                if if_exists == "replace":
                    conn.execute(
                        text(f"DROP TABLE IF EXISTS {quote_identifier(table_name)}")
                    )
                    create_table_from_headers(conn, table_name, headers)
                elif if_exists == "append":
                    pass
                else:
                    return ToolResult(
                        success=False,
                        error=ToolErrorInfo(
                            category="parameter_error",
                            message=f"Invalid if_exists value: {if_exists}",
                        ),
                    )
            else:
                create_table_from_headers(conn, table_name, headers)

            total_imported, _ = _insert_batches(conn, table_name, headers, rows, batch_size)

        return ToolResult(
            success=True,
            output=f"Imported {total_imported} row{'s' if total_imported != 1 else ''} into {table_name}",
            data={
                "table": table_name,
                "columns": headers,
                "rows_imported": total_imported,
            },
        )
    except Exception as e:
        return _to_tool_error(e)


def _import_csv_to_db_impl(
    ctx: RunContext[Settings],
    filepath: str,
    table_name: str,
    encoding: str = "utf-8",
    if_exists: str = "fail",
    batch_size: int = 100,
    row_filter: Callable[[dict[str, Any]], bool] | None = None,
    row_transform: Callable[[dict[str, Any]], dict[str, Any]] | None = None,
) -> ToolResult:
    if not os.path.isfile(filepath):
        return ToolResult(
            success=False,
            error=ToolErrorInfo(
                category="resource_not_found",
                message=f"File not found: {filepath}",
            ),
        )

    try:
        with open(filepath, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f)
            try:
                headers = [str(h) for h in next(reader)]
            except StopIteration:
                headers = []
            raw_rows = [row for row in reader]

        rows = _apply_hooks(raw_rows, headers, row_filter, row_transform)
        return _import_rows_to_db(ctx, table_name, headers, rows, if_exists, batch_size)
    except Exception as e:
        return _to_tool_error(e)


@db_tool(name="import_csv_to_db", category="write", timeout=60, requires_approval=True, validator=validate_import_args)
def import_csv_to_db(
    ctx: RunContext[Settings],
    filepath: str,
    table_name: str,
    encoding: str = "utf-8",
    if_exists: str = "fail",
    batch_size: int = 100,
) -> ToolResult:
    """Agent-visible entry point: imports a CSV file into the database."""
    return _import_csv_to_db_impl(
        ctx, filepath, table_name, encoding, if_exists, batch_size, None, None
    )


def _import_excel_to_db_impl(
    ctx: RunContext[Settings],
    filepath: str,
    table_name: str,
    sheet_name: str | int = 0,
    header_row: int = 1,
    if_exists: str = "fail",
    batch_size: int = 100,
    row_filter: Callable[[dict[str, Any]], bool] | None = None,
    row_transform: Callable[[dict[str, Any]], dict[str, Any]] | None = None,
) -> ToolResult:
    if not os.path.isfile(filepath):
        return ToolResult(
            success=False,
            error=ToolErrorInfo(
                category="resource_not_found",
                message=f"File not found: {filepath}",
            ),
        )

    try:
        from openpyxl import load_workbook
    except ImportError as err:
        return ToolResult(
            success=False,
            error=ToolErrorInfo(category="parameter_error", message=str(err)),
        )

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if isinstance(sheet_name, int):
            if sheet_name < 0 or sheet_name >= len(wb.worksheets):
                return ToolResult(
                    success=False,
                    error=ToolErrorInfo(
                        category="parameter_error",
                        message=f"Sheet index {sheet_name} out of range",
                    ),
                )
            ws = wb.worksheets[sheet_name]
        else:
            if sheet_name not in wb.sheetnames:
                return ToolResult(
                    success=False,
                    error=ToolErrorInfo(
                        category="parameter_error",
                        message=f"Sheet '{sheet_name}' not found",
                    ),
                )
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(min_row=header_row, values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            headers = []
        raw_rows = [list(row) for row in rows_iter]
        wb.close()

        rows = _apply_hooks(raw_rows, headers, row_filter, row_transform)
        return _import_rows_to_db(ctx, table_name, headers, rows, if_exists, batch_size)
    except Exception as e:
        return _to_tool_error(e)


@db_tool(name="import_excel_to_db", category="write", timeout=60, requires_approval=True, validator=validate_import_args)
def import_excel_to_db(
    ctx: RunContext[Settings],
    filepath: str,
    table_name: str,
    sheet_name: str | int = 0,
    header_row: int = 1,
    if_exists: str = "fail",
    batch_size: int = 100,
) -> ToolResult:
    """Agent-visible entry point: imports an Excel file into the database."""
    return _import_excel_to_db_impl(
        ctx,
        filepath,
        table_name,
        sheet_name,
        header_row,
        if_exists,
        batch_size,
        None,
        None,
    )
